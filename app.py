from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io, os, tempfile, requests

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB

# ── 格式常數（與範例.xlsx完全一致）
FONT_H     = Font(name='新細明體', size=12, bold=True)
FONT_D     = Font(name='新細明體', size=12, bold=False)
A_CTR      = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_VCT      = Alignment(vertical='center', wrap_text=True)
THIN       = Side(border_style='thin', color='000000')
B_ALL      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
COL_WIDTHS = {'A':21.375,'B':11.625,'C':10.625,'D':6.0,
              'G':11.5,'H':8.0,'I':8.75,'J':7.375,'K':17.125}
A_COL_EMU  = 1620366
ROW_NORMAL = 16.5
ROW_LAST   = 87.75
ROW_HEADER = 33.0

REQUIRED_COLS = ['商品編號','款號','尺寸','顏色名稱','商品週數','週數日期',
                 '商品中分類','TW_1(官網)級別','TW_1(官網)頭單','特別要求']

def sc(ws, coord, val, font=None, align=None, border=None):
    c = ws[coord]
    if val is not None: c.value = val
    if font:   c.font      = font
    if align:  c.alignment = align
    if border: c.border    = border

def parse_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {}
    for i, h in enumerate(headers):
        if h is not None:
            col[str(h).strip()] = i + 1

    missing = [k for k in REQUIRED_COLS if k not in col]
    if missing:
        found_sample = list(col.keys())[:12]
        raise ValueError(
            f'缺少必要欄位：{", ".join(missing)}\n'
            f'檔案中偵測到的欄位（前12個）：{found_sample}'
        )

    products, seen_ids, current = [], [], None
    for r in range(2, ws.max_row + 1):
        prod_id = str(ws.cell(r, col['商品編號']).value or '').strip()
        if not prod_id:
            continue
        style    = str(ws.cell(r, col['款號']).value or '').strip()
        size     = str(ws.cell(r, col['尺寸']).value or '').strip()
        color    = str(ws.cell(r, col['顏色名稱']).value or '').strip()
        week     = str(ws.cell(r, col['商品週數']).value or '').strip()
        date_v   = ws.cell(r, col['週數日期']).value
        date_s   = date_v.strftime('%Y-%m-%d') if hasattr(date_v, 'strftime') else str(date_v)
        category = str(ws.cell(r, col['商品中分類']).value or '').strip()
        grade    = str(ws.cell(r, col['TW_1(官網)級別']).value or '').strip()
        qty      = ws.cell(r, col['TW_1(官網)頭單']).value
        sp_raw   = str(ws.cell(r, col['特別要求']).value or '')
        special  = sp_raw[:sp_raw.index('洗滌注意')].strip() if '洗滌注意' in sp_raw else sp_raw.strip()

        # 自動抓取圖片 URL（從「商品內部圖片」欄）
        img_url = ''
        if '商品內部圖片' in col:
            img_url = str(ws.cell(r, col['商品內部圖片']).value or '').strip()

        if prod_id not in seen_ids:
            seen_ids.append(prod_id)
            current = {
                'prod_id': prod_id, 'style': style, 'week': week,
                'date': date_s, 'category': category, 'grade': grade,
                'special': special, 'img_url': img_url, 'colors': {}
            }
            products.append(current)

        if color not in current['colors']:
            current['colors'][color] = []
        current['colors'][color].append((size, int(qty) if qty else 0))

    if not products:
        raise ValueError('Excel 內沒有找到任何商品資料，請確認第一列為標題列')
    return products

def fetch_images(products):
    """自動從 img_url 下載商品圖片，回傳 {style: bytes}"""
    images_map = {}
    session = requests.Session()
    session.headers.update({'User-Agent': 'Mozilla/5.0'})
    for p in products:
        url = p.get('img_url', '')
        if not url:
            continue
        try:
            resp = session.get(url, timeout=15)
            resp.raise_for_status()
            pil_img = PILImage.open(io.BytesIO(resp.content)).convert('RGB')
            buf = io.BytesIO()
            pil_img.save(buf, format='PNG')
            images_map[p['style']] = buf.getvalue()
        except Exception:
            pass  # 下載失敗跳過，該商品圖片欄留空
    return images_map

def build_excel(products, images_map):
    wb = Workbook()
    ws = wb.active
    ws.title = '商品上架結構表(完整) (4)'

    for col_l, w in COL_WIDTHS.items():
        ws.column_dimensions[col_l].width = w

    ws.row_dimensions[1].height = ROW_HEADER
    for idx, h in enumerate(['商品圖','商品編號','款號','尺寸','顏色','週別',
                              '上架日期','中分類','官網頭單','網路級別','特別要求'], 1):
        sc(ws, f'{get_column_letter(idx)}1', h, FONT_H, A_CTR, B_ALL)

    current_row = 2
    for p in products:
        total = sum(len(s) for s in p['colors'].values())
        sr, er = current_row, current_row + total - 1

        for r in range(sr, er):
            ws.row_dimensions[r].height = ROW_NORMAL
        ws.row_dimensions[er].height = ROW_LAST

        for col_l in ['A','B','C','F','G','H','J','K']:
            ws.merge_cells(f'{col_l}{sr}:{col_l}{er}')
            for r in range(sr, er + 1):
                ws.cell(r, ord(col_l) - 64).border = B_ALL

        sc(ws, f'B{sr}', p['prod_id'],               FONT_D, A_VCT)
        sc(ws, f'C{sr}', p['style'],                  FONT_D, A_VCT)
        sc(ws, f'F{sr}', p['week'],                   FONT_D, A_VCT)
        sc(ws, f'G{sr}', f' \n \n{p["date"]}\n\n\n', FONT_D, A_VCT)
        sc(ws, f'H{sr}', p['category'],               FONT_D, A_VCT)
        sc(ws, f'J{sr}', p['grade'],                  FONT_D, A_CTR)
        sc(ws, f'K{sr}', p['special'] or None,        FONT_D, A_VCT)

        row_ptr = sr
        for color, sizes in p['colors'].items():
            cs, ce = row_ptr, row_ptr + len(sizes) - 1
            if len(sizes) > 1:
                ws.merge_cells(f'E{cs}:E{ce}')
                for r in range(cs, ce + 1):
                    ws.cell(r, 5).border = B_ALL
            sc(ws, f'E{cs}', color, FONT_D, A_CTR, B_ALL)
            for s_idx, (size, qty) in enumerate(sizes):
                r = row_ptr + s_idx
                sc(ws, f'D{r}', size,     FONT_D, A_VCT, B_ALL)
                sc(ws, f'I{r}', str(qty), FONT_D, A_VCT, B_ALL)
            row_ptr += len(sizes)

        if p['style'] in images_map:
            img_buf = io.BytesIO(images_map[p['style']])
            xl_img  = XLImage(img_buf)
            anchor  = TwoCellAnchor()
            anchor.editAs = 'twoCell'
            anchor._from  = AnchorMarker(col=0, colOff=0,         row=sr-1, rowOff=0)
            anchor.to     = AnchorMarker(col=0, colOff=A_COL_EMU, row=er,   rowOff=1000)
            xl_img.anchor = anchor
            ws.add_image(xl_img)

        current_row = er + 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ── Routes ────────────────────────────────────────────────────

@app.route('/api/proxy-image')
def proxy_image():
    """讓伺服器代替瀏覽器去抓圖片，解決 CORS 問題"""
    url = request.args.get('url', '')
    if not url:
        return '', 400
    try:
        resp = requests.get(url, timeout=15, headers={'User-Agent': 'Mozilla/5.0'})
        resp.raise_for_status()
        return resp.content, 200, {
            'Content-Type': resp.headers.get('Content-Type', 'image/png'),
            'Access-Control-Allow-Origin': '*'
        }
    except Exception as e:
        return str(e), 500

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/preview', methods=['POST'])
def preview():
    if 'excel' not in request.files:
        return jsonify({'error': '請選擇 Excel 檔案後再上傳'}), 400
    f = request.files['excel']
    if not f.filename:
        return jsonify({'error': '檔案名稱為空，請重新選擇'}), 400

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    f.save(tmp.name)
    tmp.close()
    try:
        products = parse_excel(tmp.name)
        result = []
        for p in products:
            total = sum(len(s) for s in p['colors'].values())
            result.append({
                'prod_id':    p['prod_id'],
                'style':      p['style'],
                'category':   p['category'],
                'week':       p['week'],
                'date':       p['date'],
                'grade':      p['grade'],
                'special':    p['special'],
                'img_url':    p['img_url'],
                'colors':     list(p['colors'].keys()),
                'total_rows': total,
            })
        return jsonify({'products': result, 'count': len(result)})
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try: os.unlink(tmp.name)
        except: pass

@app.route('/api/generate', methods=['POST'])
def generate():
    if 'excel' not in request.files:
        return jsonify({'error': '請選擇 Excel 檔案'}), 400

    f = request.files['excel']
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    f.save(tmp.name)
    tmp.close()
    try:
        products = parse_excel(tmp.name)
    except Exception as e:
        try: os.unlink(tmp.name)
        except: pass
        return jsonify({'error': str(e)}), 400
    finally:
        try: os.unlink(tmp.name)
        except: pass

    # 優先使用瀏覽器傳來的圖片，若沒有則嘗試從 URL 下載
    images_map = {}
    all_keys = [k for k in request.files.keys()]
    print(f"[DEBUG] 收到的檔案 keys: {all_keys}")
    for key, img_file in request.files.items():
        if key == 'excel':
            continue
        try:
            pil_img = PILImage.open(img_file.stream).convert('RGB')
            buf = io.BytesIO()
            pil_img.save(buf, format='PNG')
            style_name = key.replace('img_', '', 1)
            images_map[style_name] = buf.getvalue()
            print(f"[DEBUG] 成功處理圖片: {style_name}")
        except Exception as e:
            print(f"[DEBUG] 圖片處理失敗 {key}: {e}")

    print(f"[DEBUG] 總共處理圖片數: {len(images_map)}")

    # 沒收到圖片才嘗試從 URL 下載（備用）
    if not images_map:
        print("[DEBUG] 嘗試從 URL 下載圖片")
        images_map = fetch_images(products)
    print(f"[DEBUG] 最終圖片數: {len(images_map)}")

    try:
        out_buf = build_excel(products, images_map)
        week = products[0]['week'] if products else 'output'
        filename = f'{week}_商品上架結構表.xlsx'
        return send_file(
            out_buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': f'產生失敗：{str(e)}'}), 500

if __name__ == '__main__':
    print('\n' + '='*50)
    print('  商品上架結構表產生器 已啟動')
    print('  請用瀏覽器開啟：http://localhost:5000')
    print('='*50 + '\n')
    app.run(debug=False, host='0.0.0.0', port=5000)
